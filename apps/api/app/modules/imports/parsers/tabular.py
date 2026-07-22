import csv
from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.modules.content.account_models import Platform
from app.modules.metrics.definitions import validate_metric_values
from app.modules.metrics.models import ContentType


@dataclass(frozen=True)
class ParsedTable:
    headers: list[str]
    rows: list[dict[str, object]]


COMMON_ALIASES: dict[str, str] = {
    "作品链接": "work_url",
    "笔记链接": "work_url",
    "链接": "work_url",
    "标题": "title",
    "文案": "body",
    "正文": "body",
    "发布时间": "published_at",
    "数据时间": "collected_at",
    "采集时间": "collected_at",
    "内容类型": "content_type",
}

PLATFORM_ALIASES: dict[Platform, dict[str, str]] = {
    Platform.DOUYIN: {
        "作品ID": "platform_content_id",
        "播放量": "metric.views",
        "点赞数": "metric.likes",
        "评论数": "metric.comments",
        "分享数": "metric.shares",
        "收藏数": "metric.favorites",
        "5秒完播率": "metric.completion_rate_5s",
        "完播率": "metric.completion_rate",
        "平均播放时长": "metric.average_watch_duration",
        "新增粉丝": "metric.followers_gained",
    },
    Platform.XIAOHONGSHU: {
        "笔记ID": "platform_content_id",
        "曝光量": "metric.impressions",
        "阅读/播放量": "metric.views",
        "阅读量": "metric.views",
        "点赞": "metric.likes",
        "评论": "metric.comments",
        "分享": "metric.shares",
        "收藏": "metric.favorites",
        "封面点击率": "metric.cover_click_rate",
        "新增粉丝": "metric.followers_gained",
    },
}


def _json_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def read_tabular(file_name: str, data: bytes) -> ParsedTable:
    suffix = file_name.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        text = data.decode("utf-8-sig")
        reader = csv.reader(StringIO(text))
        records = list(reader)
    elif suffix == "xlsx":
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        records = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
    else:
        raise ValueError("only .csv and .xlsx files are supported")
    if not records:
        raise ValueError("import file is empty")
    headers = [str(value or "").strip() for value in records[0]]
    if not any(headers):
        raise ValueError("import file has no headers")
    rows = [
        {
            header: _json_value(value)
            for header, value in zip(headers, record, strict=False)
            if header
        }
        for record in records[1:]
        if any(value not in (None, "") for value in record)
    ]
    return ParsedTable(headers=headers, rows=rows)


def suggest_headers(headers: list[str], platform: Platform) -> list[dict[str, object]]:
    aliases = {**COMMON_ALIASES, **PLATFORM_ALIASES[platform]}
    return [
        {
            "source_header": header,
            "target_field": aliases.get(header),
            "confidence": 1.0 if header in aliases else 0.0,
            "high_confidence": header in aliases,
        }
        for header in headers
    ]


def _parse_datetime(value: object) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = str(value or "").strip()
        if not text:
            raise ValueError("value is required")
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            for pattern in ("%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y/%m/%d"):
                try:
                    parsed = datetime.strptime(text, pattern)
                    break
                except ValueError:
                    continue
            else:
                raise ValueError("invalid datetime") from None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _parse_number(value: object) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError("invalid number")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace(" ", "")
    if not text:
        raise ValueError("invalid number")
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    multiplier = Decimal(1)
    if text.endswith("万"):
        text = text[:-1]
        multiplier = Decimal(10_000)
    elif text.endswith("亿"):
        text = text[:-1]
        multiplier = Decimal(100_000_000)
    try:
        parsed = Decimal(text) * multiplier
    except InvalidOperation:
        raise ValueError("invalid number") from None
    return parsed / 100 if percent else parsed


def _mapping_dict(header_mappings: list[dict[str, object]]) -> dict[str, str]:
    return {
        str(item["source_header"]): str(item["target_field"])
        for item in header_mappings
        if item.get("target_field")
    }


def normalize_tabular_row(
    raw_data: dict[str, object],
    header_mappings: list[dict[str, object]],
    platform: Platform,
    content_type: ContentType,
) -> tuple[dict[str, object], list[dict[str, str]]]:
    mapped: dict[str, object] = {}
    metrics: dict[str, str] = {}
    errors: list[dict[str, str]] = []
    for source, target in _mapping_dict(header_mappings).items():
        value = raw_data.get(source)
        if target.startswith("metric."):
            key = target.removeprefix("metric.")
            if value in (None, ""):
                continue
            try:
                metrics[key] = str(_parse_number(value))
            except ValueError as error:
                errors.append({"field": target, "message": str(error)})
        else:
            mapped[target] = value
    mapped["metrics"] = metrics
    return normalize_manual_row(mapped, platform, content_type, errors=errors)


def normalize_manual_row(
    raw_data: dict[str, object],
    platform: Platform,
    content_type: ContentType,
    *,
    errors: list[dict[str, str]] | None = None,
) -> tuple[dict[str, object], list[dict[str, str]]]:
    row_errors = list(errors or [])
    normalized: dict[str, object] = {}
    for field in ("platform_content_id", "work_url", "title", "body"):
        value = raw_data.get(field)
        normalized[field] = str(value).strip() if value not in (None, "") else None
    if not normalized["title"]:
        row_errors.append({"field": "title", "message": "title is required"})
    normalized["body"] = normalized["body"] or ""
    for field in ("published_at", "collected_at"):
        try:
            normalized[field] = _parse_datetime(raw_data.get(field)).isoformat()
        except ValueError as error:
            normalized[field] = None
            row_errors.append({"field": field, "message": str(error)})
    if normalized["published_at"] and normalized["collected_at"]:
        if datetime.fromisoformat(str(normalized["collected_at"])) < datetime.fromisoformat(
            str(normalized["published_at"])
        ):
            row_errors.append(
                {
                    "field": "collected_at",
                    "message": "collection time cannot be before publication",
                }
            )

    metric_values = raw_data.get("metrics")
    normalized_metrics: dict[str, str] = {}
    if isinstance(metric_values, dict):
        for key, value in metric_values.items():
            if value in (None, ""):
                continue
            try:
                normalized_metrics[str(key)] = str(_parse_number(value))
            except ValueError as error:
                row_errors.append(
                    {"field": f"metric.{key}", "message": str(error)}
                )
    if not normalized_metrics:
        row_errors.append(
            {"field": "metrics", "message": "at least one metric is required"}
        )
    else:
        try:
            validate_metric_values(platform, content_type, normalized_metrics)
        except ValueError as error:
            row_errors.append({"field": "metrics", "message": str(error)})
    normalized["metrics"] = normalized_metrics
    return normalized, row_errors
